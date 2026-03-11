"""
╔══════════════════════════════════════════════════════════════════╗
║     毕业论文数据抓取脚本 v3.0  Final                             ║
║     框架: F1经济利益 F2地缘政治 F3中国因素 F4制度合规 F5谈判行为  ║
║     变量: 2Y + 21IV = 23个 | 国家: 106个                         ║
║                                                                  ║
║     pip install requests pandas openpyxl tqdm                   ║
╚══════════════════════════════════════════════════════════════════╝
"""

import requests, pandas as pd, time, json, os, shutil
from openpyxl import load_workbook

# ══════════════════════════════════════════════════════════════════
#  CONFIG — 填写你的 API Keys
# ══════════════════════════════════════════════════════════════════
GTA_KEY        = "4c6c504904a611e843b6ef0a361d64f6e740d49f"   # ✅ 已有
COMTRADE_KEY   = ""   # 申请: comtradeplus.un.org → 注册免费500次/天

EXCEL_TEMPLATE = "Thesis_Final_Dataset_v3.xlsx"
OUTPUT_FILE    = "Thesis_Data_Filled_v3.xlsx"
YEAR           = 2024   # 自变量用2024年数据；Y用2025白宫Annex I
SLEEP          = 0.8

# ══════════════════════════════════════════════════════════════════
#  国家列表
# ══════════════════════════════════════════════════════════════════
COUNTRIES = [
    ("AFG","004"),("ALB","008"),("DZA","012"),("AGO","024"),("ARG","032"),
    ("ARM","051"),("AUS","036"),("AUT","040"),("AZE","031"),("BGD","050"),
    ("BLR","112"),("BEL","056"),("BOL","068"),("BRA","076"),("BGR","100"),
    ("KHM","116"),("CAN","124"),("CHL","152"),("CHN","156"),("COL","170"),
    ("COD","180"),("CRI","188"),("HRV","191"),("CZE","203"),("DNK","208"),
    ("DOM","214"),("ECU","218"),("EGY","818"),("SLV","222"),("ETH","231"),
    ("FIN","246"),("FRA","250"),("GEO","268"),("DEU","276"),("GHA","288"),
    ("GRC","300"),("GTM","320"),("HND","340"),("HKG","344"),("HUN","348"),
    ("IND","356"),("IDN","360"),("IRN","364"),("IRL","372"),("ISR","376"),
    ("ITA","380"),("JPN","392"),("JOR","400"),("KAZ","398"),("KEN","404"),
    ("KOR","410"),("KWT","414"),("LAO","418"),("LBN","422"),("LTU","440"),
    ("LUX","442"),("MYS","458"),("MLT","470"),("MEX","484"),("MDA","498"),
    ("MNG","496"),("MAR","504"),("MOZ","508"),("MMR","104"),("NPL","524"),
    ("NLD","528"),("NZL","554"),("NGA","566"),("NOR","578"),("OMN","512"),
    ("PAK","586"),("PAN","591"),("PRY","600"),("PER","604"),("PHL","608"),
    ("POL","616"),("PRT","620"),("QAT","634"),("ROU","642"),("RUS","643"),
    ("SAU","682"),("SEN","686"),("SRB","688"),("SGP","702"),("SVK","703"),
    ("SVN","705"),("ZAF","710"),("ESP","724"),("LKA","144"),("SWE","752"),
    ("CHE","756"),("TWN","158"),("TZA","834"),("THA","764"),("TUN","788"),
    ("TUR","792"),("UGA","800"),("UKR","804"),("ARE","784"),("GBR","826"),
    ("URY","858"),("UZB","860"),("VEN","862"),("VNM","704"),("ZMB","894"),
    ("ZWE","716"),
]
ISO3_LIST = [c[0] for c in COUNTRIES]
M49_MAP   = {c[0]: c[1] for c in COUNTRIES}

print("="*60)
print("📚 毕业论文数据抓取 v3.0")
print(f"   国家数: {len(COUNTRIES)} | 目标年份: {YEAR}")
print("="*60)

results = {iso3: {} for iso3 in ISO3_LIST}

def safe_get(url, params=None, headers=None, timeout=20, label=""):
    try:
        r = requests.get(url, params=params, headers=headers, timeout=timeout)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"  ⚠ {label}: {e}")
        return None

# ══════════════════════════════════════════════════════════════════
#  F1-A: World Bank API — 制造业出口占比 + GDP（无需Key）
# ══════════════════════════════════════════════════════════════════
print("\n[1/7] 🌐 World Bank API (F1: 制造业出口, GDP)")

WB_VARS = {
    "TX.VAL.MANF.ZS.UN": "F1_ManufShare",   # 制造业出口/总出口%
    "NY.GDP.MKTP.CD":    "WB_GDP",           # GDP现价美元
    "SP.POP.TOTL":       "WB_Pop",           # 人口（F2控制变量）
}

for ind, col in WB_VARS.items():
    url = (f"https://api.worldbank.org/v2/country/all/indicator/{ind}"
           f"?format=json&date={YEAR}&per_page=300&mrv=1")
    data = safe_get(url, label=f"WB/{ind}")
    if data and len(data) > 1 and data[1]:
        n = 0
        for row in data[1]:
            iso3 = row.get("countryiso3code","")
            val  = row.get("value")
            if iso3 in results and val is not None:
                results[iso3][col] = round(float(val), 4)
                n += 1
        print(f"  ✓ {col}: {n} 国")
    time.sleep(SLEEP)

# ══════════════════════════════════════════════════════════════════
#  F1-B: UN Comtrade — 对美顺差、对美出口依赖、HS品类、对华依赖（需Key）
# ══════════════════════════════════════════════════════════════════
print("\n[2/7] 📦 UN Comtrade API (F1: 贸易顺差/依赖, F3: 对华依赖)")

if not COMTRADE_KEY:
    print("  ⚠ 未填写 COMTRADE_KEY — 跳过 Comtrade 数据")
    print("  → 请前往 comtradeplus.un.org 注册获取免费 Key")
else:
    BASE = "https://comtradeapi.un.org/data/v1/get/C/A/HS"

    def ct_fetch(reporter, partner, flow):
        p = {
            "reporterCode": reporter,
            "partnerCode": partner,
            "period": str(YEAR),
            "cmdCode": "TOTAL",
            "flowCode": flow,
            "maxRecords": 1,
            "format": "JSON",
            "subscription-key": COMTRADE_KEY,
        }
        d = safe_get(BASE, params=p, label=f"CT {reporter} {flow}→{partner}")
        if d and d.get("data"):
            return float(d["data"][0].get("primaryValue", 0) or 0)
        return None

    def ct_top_hs(reporter):
        """获取对美最大出口HS2品类并归类"""
        p = {
            "reporterCode": reporter,
            "partnerCode": "842",
            "period": str(YEAR),
            "cmdCode": "AG2",
            "flowCode": "X",
            "maxRecords": 5,
            "format": "JSON",
            "subscription-key": COMTRADE_KEY,
        }
        d = safe_get(BASE, params=p, label=f"CT HS {reporter}")
        if d and d.get("data"):
            rows = sorted(d["data"],
                          key=lambda x: x.get("primaryValue",0), reverse=True)
            if rows:
                hs = int(str(rows[0].get("cmdCode","0"))[:2])
                # 分类: 1=制造 2=农业 3=资源 4=服务/其他
                if hs in range(25,98):    return 1   # 制造业
                elif hs in range(1,25):   return 2   # 农业/食品
                elif hs in range(25,28):  return 3   # 矿产资源
                else:                     return 4
        return None

    print("  抓取各国对美/对华贸易数据...")
    for n, iso3 in enumerate(ISO3_LIST, 1):
        m49 = M49_MAP.get(iso3)
        if not m49:
            continue

        # 对美出口
        exp_us = ct_fetch(m49, "842", "X")
        time.sleep(SLEEP)
        imp_us = ct_fetch(m49, "842", "M")
        time.sleep(SLEEP)

        if exp_us is not None:
            results[iso3]["F1_ExportToUS"] = round(exp_us/1e9, 3)
        if exp_us is not None and imp_us is not None:
            results[iso3]["F1_Surplus"]    = round((exp_us - imp_us)/1e9, 3)

        # 对美出口依赖 = 对美出口 / 总出口
        exp_wld = ct_fetch(m49, "0", "X")  # 0=World
        time.sleep(SLEEP)
        if exp_us and exp_wld and exp_wld > 0:
            results[iso3]["F1_ExpDep"] = round(exp_us/exp_wld*100, 2)

        # 对华出口依赖
        exp_cn = ct_fetch(m49, "156", "X")
        time.sleep(SLEEP)
        if exp_cn and exp_wld and exp_wld > 0:
            results[iso3]["F3_ChinaDep"] = round(exp_cn/exp_wld*100, 2)

        # HS品类
        hs_cat = ct_top_hs(m49)
        if hs_cat:
            results[iso3]["F1_HSCode"] = hs_cat
        time.sleep(SLEEP)

        if n % 10 == 0:
            print(f"  进度 {n}/{len(ISO3_LIST)}")

# ══════════════════════════════════════════════════════════════════
#  F1-C: Senate LDA + FARA — 游说支出（无需Key）
# ══════════════════════════════════════════════════════════════════
print("\n[3/7] 🏛️  Senate LDA + FARA 游说数据 (F1: 游说支出)")

# 各国政府/企业关键词
LOBBY_TERMS = {
    "CHN": ["China","CNOOC","Huawei","Alibaba","ByteDance","COSCO"],
    "JPN": ["Japan","Toyota","Sony","SoftBank","Mitsubishi","Honda"],
    "DEU": ["Germany","Volkswagen","BMW","Siemens","BASF","Bayer"],
    "KOR": ["Korea","Samsung","LG","Hyundai","SK Group","Posco"],
    "IND": ["India","Tata","Infosys","Wipro","Reliance"],
    "MEX": ["Mexico","PEMEX","Cemex","America Movil"],
    "CAN": ["Canada","Bombardier","Barrick","Enbridge","TC Energy"],
    "GBR": ["Britain","United Kingdom","BP","Shell","HSBC","BAE Systems"],
    "FRA": ["France","Total","Airbus","LVMH","BNP Paribas","Sanofi"],
    "ITA": ["Italy","ENI","Fiat","Leonardo"],
    "BRA": ["Brazil","Petrobras","Vale","Embraer"],
    "AUS": ["Australia","BHP","Rio Tinto","Macquarie"],
    "SAU": ["Saudi","Aramco","SABIC"],
    "ARE": ["Emirates","UAE","ADNOC","DP World","Mubadala"],
    "ISR": ["Israel","AIPAC","Rafael","Elbit"],
    "VNM": ["Vietnam"],
    "TWN": ["Taiwan","TSMC","Foxconn","MediaTek"],
    "CHE": ["Switzerland","Nestle","Novartis","Roche","UBS"],
    "SGP": ["Singapore","Temasek","DBS"],
    "NLD": ["Netherlands","ASML","ING","Heineken"],
    "SWE": ["Sweden","Ericsson","Volvo","Spotify"],
    "NOR": ["Norway","Equinor"],
    "ZAF": ["South Africa","Anglo American"],
    "TUR": ["Turkey","Koc Group"],
    "QAT": ["Qatar","Qatar Airways"],
    "KWT": ["Kuwait"],
    "RUS": ["Russia"],
    "IDN": ["Indonesia","Pertamina"],
    "MYS": ["Malaysia","Petronas"],
    "THA": ["Thailand"],
    "PHL": ["Philippines"],
}

def lda_search(term):
    """LDA 官方 API 搜索游说金额"""
    total = 0.0
    params = {
        "client_name": term,
        "filing_year":f"2023,2024",
        "page_size": 50,
        "format": "json",
    }
    data = safe_get("https://lda.senate.gov/api/v1/filings/",
                    params=params, label=f"LDA/{term}")
    if data:
        for f in data.get("results", []):
            amt = f.get("income") or f.get("expenses") or 0
            try:    total += float(str(amt).replace(",",""))
            except: pass
    return total

def fara_count(term):
    """FARA DOJ API 外国代理登记数"""
    data = safe_get(
        "https://efile.fara.gov/api/v1/ForeignPrincipal/json",
        label="FARA")
    if not data:
        return 0
    count = 0
    principals = data.get("ForeignPrincipals",{}).get("ForeignPrincipal",[])
    for p in principals:
        country = p.get("Country","").upper()
        if term.upper() in country:
            count += 1
    return count

for iso3, terms in LOBBY_TERMS.items():
    if iso3 not in results:
        continue
    total_usd = 0.0
    for term in terms[:3]:  # 取前3个关键词避免超时
        total_usd += lda_search(term)
        time.sleep(0.4)
    results[iso3]["F1_Lobby"] = round(total_usd / 1e6, 3)

print(f"  ✓ 游说数据已处理 {len(LOBBY_TERMS)} 个国家")

# FARA（一次性全量获取，按国家名匹配）
FARA_MAP = {
    "CHN":"CHINA","JPN":"JAPAN","DEU":"GERMANY","KOR":"SOUTH KOREA",
    "GBR":"UNITED KINGDOM","FRA":"FRANCE","CAN":"CANADA","AUS":"AUSTRALIA",
    "SAU":"SAUDI ARABIA","ISR":"ISRAEL","IND":"INDIA","RUS":"RUSSIA",
    "TWN":"TAIWAN","MEX":"MEXICO","BRA":"BRAZIL","TUR":"TURKEY",
    "ARE":"UAE","QAT":"QATAR","SGP":"SINGAPORE","CHE":"SWITZERLAND",
    "NLD":"NETHERLANDS","SWE":"SWEDEN","JOR":"JORDAN",
}
fara_all = safe_get("https://efile.fara.gov/api/v1/ForeignPrincipal/json",
                    label="FARA全量")
if fara_all:
    from collections import Counter
    country_cnt = Counter()
    for p in fara_all.get("ForeignPrincipals",{}).get("ForeignPrincipal",[]):
        c = p.get("Country","").upper().strip()
        if c: country_cnt[c] += 1
    for iso3, fname in FARA_MAP.items():
        if iso3 in results:
            results[iso3]["F1_FARA"] = country_cnt.get(fname, 0)
    print(f"  ✓ FARA: {len(country_cnt)} 个国家/地区有记录")

# ══════════════════════════════════════════════════════════════════
#  F2: 地理距离 + 静态编码（盟友、对手）
# ══════════════════════════════════════════════════════════════════
print("\n[4/7] 🌍 地缘政治静态数据 (F2)")

# 距华盛顿DC直线距离 (km)
DIST_DC = {
    "AFG":11232,"ALB":8558,"DZA":7040,"AGO":10971,"ARG":9085,"ARM":9607,
    "AUS":16233,"AUT":7338,"AZE":9405,"BGD":13026,"BLR":8013,"BEL":6450,
    "BOL":7200,"BRA":7715,"BGR":8688,"KHM":14527,"CAN":743,"CHL":8560,
    "CHN":11671,"COL":4418,"COD":10500,"CRI":3400,"HRV":8200,"CZE":7200,
    "DNK":6656,"DOM":2680,"ECU":4700,"EGY":9376,"SLV":3200,"ETH":12055,
    "FIN":7100,"FRA":6670,"GEO":9356,"DEU":6985,"GHA":9330,"GRC":8800,
    "GTM":3100,"HND":3300,"HKG":13100,"HUN":7800,"IND":12507,"IDN":16346,
    "IRN":10700,"IRL":5757,"ISR":9420,"ITA":8116,"JPN":10838,"JOR":9800,
    "KAZ":9800,"KEN":12553,"KOR":11081,"KWT":10400,"LAO":14200,"LBN":9600,
    "LTU":7700,"LUX":6500,"MYS":15289,"MLT":8200,"MEX":3068,"MDA":8300,
    "MNG":10700,"MAR":6534,"MOZ":13898,"MMR":14400,"NPL":12100,"NLD":6230,
    "NZL":14544,"NGA":9521,"NOR":6473,"OMN":11700,"PAK":11660,"PAN":3500,
    "PRY":7900,"PER":5400,"PHL":14271,"POL":7298,"PRT":5800,"QAT":10900,
    "ROU":8500,"RUS":9167,"SAU":10636,"SEN":7700,"SRB":8400,"SGP":15345,
    "SVK":7500,"SVN":7900,"ZAF":14366,"ESP":6070,"LKA":14400,"SWE":6875,
    "CHE":7173,"TWN":12199,"TZA":13119,"THA":14117,"TUN":8100,"TUR":9094,
    "UGA":12000,"UKR":9103,"ARE":11200,"GBR":5914,"URY":8600,"UZB":10500,
    "VEN":3900,"VNM":13524,"ZMB":13319,"ZWE":13000,
}

# 美国条约盟友（NATO + 亚太条约盟友）
US_ALLIES = {
    "ALB","BEL","BGR","CAN","HRV","CZE","DNK","EST","FIN","FRA",
    "DEU","GRC","HUN","ISL","ITA","LVA","LTU","LUX","MNE","NLD",
    "MKD","NOR","POL","PRT","ROU","SVK","SVN","ESP","SWE","TUR",
    "GBR",
    # 亚太条约盟友
    "JPN","KOR","AUS","NZL","PHL","THA",
    "ISR",
}

for iso3 in ISO3_LIST:
    results[iso3]["F2_Dist"] = DIST_DC.get(iso3)
    results[iso3]["F2_Ally"] = 1 if iso3 in US_ALLIES else 0

print(f"  ✓ 地理距离: {sum(1 for v in results.values() if v.get('F2_Dist'))} 国")
print(f"  ✓ 盟友编码: {sum(v.get('F2_Ally',0) for v in results.values())} 个盟友")

# ══════════════════════════════════════════════════════════════════
#  F3: RCEP + BRI 静态编码
# ══════════════════════════════════════════════════════════════════
print("\n[5/7] 🐉 中国因素静态数据 (F3: RCEP/BRI)")

RCEP = {"AUS","BRN","KHM","CHN","IDN","JPN","KOR","LAO",
        "MYS","MMR","NZL","PHL","SGP","THA","VNM"}

BRI  = {
    "AFG","ALB","DZA","AGO","ARM","AZE","BGD","BLR","BEL","BOL",
    "BIH","BGR","KHM","CMR","CHN","COD","CRI","HRV","CZE","EGY",
    "ETH","GAB","GEO","GHA","GRC","GTM","HUN","IDN","IRN","IRQ",
    "ITA","JOR","KAZ","KEN","KGZ","LAO","LBN","LTU","MKD","MYS",
    "MLI","MLT","MRT","MDA","MNG","MAR","MOZ","MMR","NAM","NPL",
    "NGA","OMN","PAK","PAN","PER","PHL","POL","PRT","ROU","RUS",
    "SAU","SEN","SRB","SGP","LKA","TJK","TZA","THA","TUN","TUR",
    "TKM","UGA","UKR","ARE","URY","UZB","VEN","VNM","ZMB","ZWE",
}

for iso3 in ISO3_LIST:
    results[iso3]["F3_RCEP"] = 1 if iso3 in RCEP else 0
    results[iso3]["F3_BRI"]  = 1 if iso3 in BRI  else 0

print(f"  ✓ RCEP: {sum(v.get('F3_RCEP',0) for v in results.values())} 个成员")
print(f"  ✓ BRI:  {sum(v.get('F3_BRI',0) for v in results.values())} 个参与国")

# ══════════════════════════════════════════════════════════════════
#  F4: 制度合规静态编码（WTO成员已删，保留FTA/NME/301）
# ══════════════════════════════════════════════════════════════════
print("\n[6/7] 📜 制度合规静态数据 (F4: FTA/NME/301)")

US_FTA = {"AUS","BHR","CAN","CHL","COL","CRI","DOM","SLV","GTM","HND",
           "ISR","JOR","KOR","MEX","MAR","NIC","OMN","PAN","PER","SGP","GBR"}

NME    = {"CHN","VNM","KAZ","KGZ","TJK","UZB","ARM","AZE","GEO","MDA",
           "RUS","UKR","BLR","ALB"}

USTR_301 = {"CHN","IND","RUS","ARG","IDN","VEN","DZA","BGR","PAK",
             "UKR","VNM","THA","TUR","EGY","CHE","ITA","KOR","MEX"}

for iso3 in ISO3_LIST:
    results[iso3]["F4_FTA"] = 1 if iso3 in US_FTA   else 0
    results[iso3]["F4_NME"] = 1 if iso3 in NME      else 0
    results[iso3]["F4_301"] = 1 if iso3 in USTR_301 else 0

print(f"  ✓ FTA: {sum(v.get('F4_FTA',0) for v in results.values())} 个")
print(f"  ✓ NME: {sum(v.get('F4_NME',0) for v in results.values())} 个")
print(f"  ✓ 301: {sum(v.get('F4_301',0) for v in results.values())} 个")

# ══════════════════════════════════════════════════════════════════
#  F5-GTA: 报复性关税（用GTA API）
# ══════════════════════════════════════════════════════════════════
print("\n[7/7] 🛡️  GTA API — 报复性关税 (F5_Retaliate)")

gta_retaliate = {}
try:
    PAGES = 10
    for page in range(1, PAGES + 1):
        params = {
            "key":   GTA_KEY,
            "limit": 500,
            "offset":(page-1)*500,
            "implementation_period_from": "2025-01-01",
            "implementation_period_to":   "2026-02-28",
            "format": "json",
        }
        data = safe_get("https://globaltradealert.org/api/v1/interventions",
                        params=params, label=f"GTA p{page}")
        if not data:
            break
        items = data if isinstance(data, list) else data.get("data", [])
        if not items:
            break
        for item in items:
            implementing = item.get("implementing_jurisdiction","")
            affected     = item.get("affected_jurisdiction","")
            # 其他国家对美国实施的报复措施
            if "840" in str(affected) or "USA" in str(affected).upper():
                gta_retaliate[implementing] = gta_retaliate.get(implementing,0)+1
        if len(items) < 500:
            break
        time.sleep(SLEEP)
    print(f"  ✓ GTA报复数据: {len(gta_retaliate)} 个实施方")
except Exception as e:
    print(f"  ⚠ GTA 失败: {e}")

for iso3 in ISO3_LIST:
    cnt = gta_retaliate.get(iso3, 0)
    results[iso3]["F5_Retaliate"] = 1 if cnt > 0 else 0

# ══════════════════════════════════════════════════════════════════
#  Y变量手动编码模板（2025年Liberation Day税率）
# ══════════════════════════════════════════════════════════════════
# 来源: White House Annex I (2025-04-02) + Yale Budget Lab
# 10%=基准税率国家；下方列出高税率国家
Y1_TARIFF = {
    "CHN": 145.0,   # 叠加后最高（含芬太尼等）
    "VNM":  46.0,
    "KHM":  49.0,
    "THA":  36.0,
    "IDN":  32.0,
    "TWN":  32.0,
    "IND":  26.0,
    "KOR":  25.0,
    "JPN":  24.0,
    "MYS":  24.0,
    "CHE":  31.0,
    "ZAF":  30.0,
    "PAK":  29.0,
    "TUN":  28.0,
    "BGD":  37.0,
    "LKA":  44.0,
    "MMR":  44.0,
    "LAO":  48.0,
    "MYS":  24.0,
    "ISR":  17.0,
    "EGY":  10.0,
    "NGA":  14.0,
    "SAU":  10.0,
    "ARE":  10.0,
    "QAT":  10.0,
    "BRA":  10.0,
    "ARG":  10.0,
    "MEX":  25.0,   # USMCA外商品
    "CAN":  25.0,   # USMCA外商品
    "RUS":  35.0,   # 特殊制裁叠加
    "GBR":  10.0,
    "DEU":  20.0,   # 欧盟
    "FRA":  20.0,
    "ITA":  20.0,
    "NLD":  20.0,
    "BEL":  20.0,
    "ESP":  20.0,
    "POL":  20.0,
    "SWE":  20.0,
    "DNK":  20.0,
    "FIN":  20.0,
    "AUT":  20.0,
    "PRT":  20.0,
    "GRC":  20.0,
    "IRL":  20.0,
    "HUN":  20.0,
    "CZE":  20.0,
    "ROU":  20.0,
    "BGR":  20.0,
    "HRV":  20.0,
    "SVK":  20.0,
    "SVN":  20.0,
    "LTU":  20.0,
    "LUX":  20.0,
    "MLT":  20.0,
    "NOR":  15.0,
    "CHE":  31.0,
}
# 其余未列出国家默认10%基准税率
DEFAULT_TARIFF = 10.0

# 是否达成双边框架协议（截至2026年2月）
Y2_DEAL = {
    "GBR": 1, "CHN": 1, "VNM": 1, "IDN": 1, "MYS": 1, "KHM": 1,
    "ARG": 1, "ECU": 1, "SLV": 1, "GTM": 1, "CHE": 1, "BGD": 1,
    "JPN": 1, "KOR": 1, "IND": 0, "CAN": 0, "MEX": 0,
}

for iso3 in ISO3_LIST:
    results[iso3]["Y1_Tariff"] = Y1_TARIFF.get(iso3, DEFAULT_TARIFF)
    results[iso3]["Y2_Deal"]   = Y2_DEAL.get(iso3, 0)

print(f"\n  ✓ Y1 关税: {len([v for v in results.values() if v.get('Y1_Tariff')])} 国")
print(f"  ✓ Y2 协议: {sum(v.get('Y2_Deal',0) for v in results.values())} 国达成协议")

# ══════════════════════════════════════════════════════════════════
#  整合 DataFrame + 写入 Excel
# ══════════════════════════════════════════════════════════════════
print(f"\n💾 整合数据并写入 {OUTPUT_FILE}...")

records = [{"ISO3": iso3, **data} for iso3, data in results.items()]
df = pd.DataFrame(records).sort_values("ISO3").reset_index(drop=True)

# 保存 CSV（Stata/R 可直接导入）
csv_out = OUTPUT_FILE.replace(".xlsx","_Master.csv")
df.to_csv(csv_out, index=False, encoding="utf-8-sig")
print(f"  ✓ CSV 已保存: {csv_out}")

# 列映射: Excel变量代码 → DataFrame列名
COL_MAP = {
    "Y1_Tariff":    "Y1_Tariff",
    "Y2_Deal":      "Y2_Deal",
    "F1_Surplus":   "F1_Surplus",
    "F1_ExpDep":    "F1_ExpDep",
    "F1_ManufShare":"F1_ManufShare",
    "F1_HSCode":    "F1_HSCode",
    "F1_Lobby":     "F1_Lobby",
    "F1_FARA":      "F1_FARA",
    "F2_Dist":      "F2_Dist",
    "F2_Ally":      "F2_Ally",
    "F3_ChinaDep":  "F3_ChinaDep",
    "F3_RCEP":      "F3_RCEP",
    "F3_BRI":       "F3_BRI",
    "F4_FTA":       "F4_FTA",
    "F4_NME":       "F4_NME",
    "F4_301":       "F4_301",
    "F5_Retaliate": "F5_Retaliate",
    "WB_GDP":       "WB_GDP",
    "WB_Pop":       "WB_Pop",
}

shutil.copy(EXCEL_TEMPLATE, OUTPUT_FILE)
wb_obj = load_workbook(OUTPUT_FILE)

def fill_sheet(ws):
    iso3_col = None
    var_col  = {}
    for cell in ws[4]:
        if cell.value == "ISO3":
            iso3_col = cell.column
        if cell.value:
            var_col[cell.value] = cell.column
    if not iso3_col:
        return 0
    filled = 0
    for row_idx in range(6, ws.max_row + 1):
        iso3 = ws.cell(row=row_idx, column=iso3_col).value
        if not iso3:
            continue
        df_row = df[df["ISO3"] == iso3]
        if df_row.empty:
            continue
        r = df_row.iloc[0]
        for vcode, dcol in COL_MAP.items():
            if vcode not in var_col:
                continue
            val = r.get(dcol)
            if pd.notna(val) if not isinstance(val, str) else val:
                ws.cell(row=row_idx, column=var_col[vcode]).value = val
                filled += 1
    return filled

total = 0
for sname in wb_obj.sheetnames:
    if sname == "Codebook":
        continue
    n = fill_sheet(wb_obj[sname])
    total += n
    print(f"  ✓ {sname}: {n} 个单元格已填充")

wb_obj.save(OUTPUT_FILE)

print(f"""
{'='*60}
🎉 完成！
   Excel: {OUTPUT_FILE}
   CSV:   {csv_out}
   总填充: {total} 个单元格
{'='*60}

📌 还需手动下载补充的数据（按优先级）:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
变量          来源                     方式
─────────────────────────────────────────
F2_Bases      sipri.org / Vine         下载CSV
F2_UNVote     Harvard Dataverse        下载CSV
F2_Polity     systemicpeace.org        下载XLS
F3_TiVA       stats.oecd.org           注册下载
F4_IPRI       internationalproper...   下载Excel
F5_Stance     手动编码                 阅读USTR/白宫声明
F5_Purchase   手动编码                 阅读白宫新闻稿
F5_Invest     手动编码                 阅读Reuters/Bloomberg
Y1补充        Yale Budget Lab 2026版   更新最新税率
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
""")